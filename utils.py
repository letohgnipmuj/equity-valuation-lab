import logging
import os
import shutil
from datetime import datetime
from typing import Any, Dict, Sequence
from urllib.parse import urlencode
import numpy as np
import pandas as pd
import requests
import yfinance as yf
from openpyxl import load_workbook
import difflib

from cache import get_cache, set_cache
from constants import FMP_CACHE_TTL_SECONDS

logger = logging.getLogger(__name__)


def pull_info(statement, stock, api_key):
    """
    Fetch financial data from FMP with caching.
    """
    cache_key = f"fmp:{statement}:{stock.upper()}"
    cached_data = get_cache(cache_key)
    if cached_data:
        # print(f"Using cached FMP data for {statement}:{stock}")
        return cached_data

    if statement == "income-statement" or statement == "balance-sheet-statement" or statement == "cash-flow-statement":
        # Use stock because ticker is yfinance object
        url = f"https://financialmodelingprep.com/stable/{statement}/{stock}?apikey={api_key}"
    elif statement == "analyst-estimates":
        url = f"https://financialmodelingprep.com/stable/analyst-estimates?symbol={stock}&period=annual&page=0&limit=10&apikey={api_key}"
    elif statement == "financial-growth":
        url = f"https://financialmodelingprep.com/stable/financial-growth?symbol={stock}&apikey={api_key}"
    else:
        return None

    try:
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            if data and isinstance(data, (list, dict)):
                set_cache(cache_key, data, ttl=FMP_CACHE_TTL_SECONDS)
            return data
        else:
            logger.error("Error fetching %s: %s", statement, response.status_code)
            return None
    except Exception as e:
        logger.error("Network error while fetching %s: %s", statement, e)
        return None


def compute_sales_to_capital(
    revenue_clean: pd.Series,
    capex_series: pd.Series,
    dep_series_clean: pd.Series,
    balance: pd.DataFrame,
    last_revenue: float,
    last_capex: float
) -> float:
    """
    Computes the marginal Sales-to-Capital ratio using a multi-year window (up to 5 transitions).
    Formula: Total Delta Revenue / Total Net Reinvestment (Capex - Depreciation).
    """
    n_hist = min(len(revenue_clean), len(
        capex_series), len(dep_series_clean), 6)
    if n_hist >= 2:
        hist_rev = revenue_clean.iloc[:n_hist].values.astype(float)
        hist_capex = capex_series.iloc[:n_hist].values.astype(float)
        hist_dep = dep_series_clean.iloc[:n_hist].values.astype(float)

        # Growth over the period
        total_delta_rev = hist_rev[0] - hist_rev[n_hist - 1]

        # Reinvestment required for that growth: years [0, n_hist-2]
        total_net_capex = np.sum(
            hist_capex[:n_hist - 1] - hist_dep[:n_hist - 1])

        if total_net_capex > 0 and total_delta_rev > 0:
            return float(total_delta_rev / total_net_capex)

    # Fallback 1: Total Sales / Net PPE (Asset Productivity)
    ppe_series = get_yf_item(balance, "net ppe")
    last_ppe = float(
        ppe_series.iloc[0]) if ppe_series is not None and not ppe_series.empty else 0
    if last_ppe > 0:
        return last_revenue / last_ppe

    # Fallback 2: Sales / Gross Capex
    return last_revenue / max(last_capex, 1.0)


def get_yf_item(statement_df: pd.DataFrame, item_name: str, cutoff: float = 0.6) -> pd.Series | None:
    """
    Robust lookup for line items in yfinance statements, with fuzzy matching
    and common synonyms. Returns a Series or None if nothing reasonable is found.
    """
    synonyms = {
        "operating income": ["Operating Income", "OperatingIncome", "EBIT", "Operating Profit", "Operating Revenue", "Total Revenue"],
        "net ppe": ["Property Plant Equipment Net", "Fixed Assets", "Net Property Plant and Equipment", "Proprety Plant Equipment", "PPE"],
        "current debt": ["Short Term Debt", "Capital Lease Obligations", "Short Long Term Debt", "Current Liabilities"],
        "depreciation": ["Depreciation Amortization", "Depreciation and Amortization", "Reconciled Depreciation"],
        "cash": ["Cash And Cash Equivalents", "Cash Cash Equivalents And Short Term Investments"],
        "revenue": ["Total Revenue", "Operating Revenue", "OperatingRevenue", "Sales"],
        "total debt": ["totalDebt", "Debt", "Total Liabilities"],
    }

    if statement_df is None or statement_df.empty:
        return None

    search_key = item_name.lower().strip()

    search_terms = [item_name]
    if search_key in synonyms:
        search_terms.extend(synonyms[search_key])

    # Exact match
    for term in search_terms:
        if term in statement_df.index:
            return statement_df.loc[term]

    all_indices = statement_df.index.tolist()

    # Fuzzy match
    for term in search_terms:
        matches = difflib.get_close_matches(
            term, all_indices, n=1, cutoff=cutoff)
        if matches:
            return statement_df.loc[matches[0]]

    # Substring fallback
    for term in search_terms:
        for index_item in all_indices:
            if term.lower() in index_item.lower():
                return statement_df.loc[index_item]

    logger.warning("Could not find any data for '%s'", item_name)
    return None


def pad_with_mean(arr: Sequence[float], target_length: int) -> np.ndarray:
    arr = np.asarray(arr, dtype=float)

    mean_value = np.nanmean(arr)
    if np.isnan(mean_value):
        raise ValueError("Array contains only NaNs — cannot compute mean.")

    arr = np.where(np.isnan(arr), mean_value, arr)

    if len(arr) < target_length:
        padding = np.full(target_length - len(arr), mean_value)
        arr = np.concatenate([arr, padding])

    return arr[:target_length]


def safe_iloc0(series):
    """Safely get the first element of a pandas Series or return None."""
    if series is not None and not series.empty:
        valid_items = series.dropna()
        if not valid_items.empty:
            return valid_items.iloc[0]
    return None


def get_weights(implied_prices, temperature=1.0):
    data = np.array(implied_prices)
    mean = np.mean(data)
    std = np.std(data)

    if std == 0:
        return np.ones(len(data)) / len(data)

    scaled_distances = -np.abs(data - mean) / (std * temperature)

    exp_scores = np.exp(scaled_distances)
    weights = exp_scores / np.sum(exp_scores)

    return weights


def get_wacc(
    ticker_symbol: str,
    income: pd.DataFrame,
    r_f: float = 0.04,
    equity_risk_premium: float = 0.05,
    tax_rate: float = 0.21,
) -> float:
    """
    Compute WACC using CAPM for cost of equity and income statement for cost of debt.
    """
    ticker = yf.Ticker(ticker_symbol)
    info = ticker.info

    market_cap = info.get("marketCap")
    total_debt = info.get("totalDebt") or 0.0
    cash = info.get("totalCash") or 0.0
    beta = info.get("beta")

    if market_cap is None or beta is None:
        raise ValueError("Insufficient market data to compute WACC.")

    long_beta = (beta - 1) / 2 + 1

    if income is not None and "Interest Expense" in income.index:
        interest_series = income.loc["Interest Expense"].dropna()
        if not interest_series.empty:
            interest_expense = float(interest_series.iloc[0])
        else:
            interest_expense = 0.05
    else:
        interest_expense = 0.05

    E = float(market_cap)
    D = max(float(total_debt) - float(cash), 0.0)

    if E + D <= 0:
        # Degenerate balance sheet, treat as all equity
        w_e, w_d = 1.0, 0.0
    else:
        w_e = E / (E + D)
        w_d = D / (E + D)

    r_e = r_f + long_beta * equity_risk_premium
    r_d_pre_tax = abs(interest_expense) / \
        float(total_debt) if total_debt else 0.05
    r_d_pre_tax = max(r_d_pre_tax, r_f)
    r_d_after_tax = r_d_pre_tax * (1 - tax_rate)

    wacc = w_e * r_e + w_d * r_d_after_tax
    return float(max(wacc, r_f + 0.01))


def revenue_growth_schedule(
    estimates: pd.DataFrame,
    terminal_growth: float,
    projection_years: int,
    income: pd.DataFrame,
) -> np.ndarray:
    """
    Build a projection-years-long revenue growth schedule.

    Strategy:
    1. Compute historical CAGR as a stable anchor.
    2. If no analyst data, linearly fade CAGR down to terminal_growth.
    3. With analyst data: clip outliers (floor -20%, cap +80%), then
       blend 70% analyst / 30% historical CAGR to dampen single-year
       anomalies (fixes NVDA year-4 negative-growth artefact).
    4. For years beyond analyst coverage, linearly interpolate from the
       last blended rate to terminal_growth (smoother than exp decay).
    """
    # ── Historical CAGR anchor ────────────────────────────────────────────
    hist_cagr = 0.08  # sensible default
    if income is not None and not income.empty and "Total Revenue" in income.index:
        hist_rev = income.loc["Total Revenue"].dropna(
        ).sort_index(ascending=False)
        if len(hist_rev) >= 2:
            r_recent = float(hist_rev.iloc[0])
            r_oldest = float(hist_rev.iloc[-1])
            n = len(hist_rev) - 1
            if r_oldest > 0 and n > 0:
                hist_cagr = float(
                    np.clip((r_recent / r_oldest) ** (1 / n) - 1, -0.20, 0.60))

    df = pd.DataFrame(estimates)

    def exp_fade(g0, g_terminal, years, k=0.1):
        t = np.arange(years)
        return g_terminal + (g0 - g_terminal) * np.exp(-k * t)

    # ── No analyst data: fade hist CAGR to TGR with a healthy minimum floor
    if df.empty or "revenueAvg" not in df.columns:
        start = float(np.clip(hist_cagr, terminal_growth, 0.60))
        fallback_min_growth = max(terminal_growth, 0.05)
        start = max(start, fallback_min_growth)
        return exp_fade(start, terminal_growth, projection_years)

    df["date"] = pd.to_datetime(df["date"])
    df["year"] = df["date"].dt.year
    df = df[df["revenueAvg"].notna()].sort_values("year")

    current_year = pd.Timestamp.today().year
    df_future = df[df["year"] >= current_year]

    if len(df_future) < 2:
        start = float(np.clip(hist_cagr, terminal_growth, 0.60))
        fallback_min_growth = max(terminal_growth, 0.05)
        start = max(start, fallback_min_growth)
        return exp_fade(start, terminal_growth, projection_years)

    rev_fwd = df_future["revenueAvg"].astype(float).values
    analyst_raw = np.clip((rev_fwd[1:] / rev_fwd[:-1]) - 1, -0.20, 0.80)

    # Blend: 70% analyst, 30% historical anchor
    analyst_blended = 0.70 * analyst_raw + 0.30 * hist_cagr

    analyst_years = analyst_blended[:projection_years].tolist()
    remaining = projection_years - len(analyst_years)

    if remaining > 0:
        last_rate = float(
            np.clip(analyst_years[-1], terminal_growth, min(hist_cagr, 0.25)))
        fade = exp_fade(last_rate, terminal_growth, remaining + 2)[1:-1]
        analyst_years.extend(fade.tolist())

    return np.array(analyst_years[:projection_years], dtype=float)


def calculate_roic(
    ebit: float,
    tax_rate: float,
    total_assets: float,
    cash: float,
    current_liabilities: float,
    short_term_debt: float = 0.0,
) -> float:
    """
    Standard ROIC calculation: NOPAT / Invested Capital.
    Invested Capital = (Total Assets - Cash) - (Current Liabilities - Short-Term Debt)
    """
    nopat = ebit * (1 - tax_rate)
    operating_assets = total_assets - cash
    # Non-interest bearing current liabilities
    nibcl = current_liabilities - short_term_debt
    invested_capital = operating_assets - nibcl

    if invested_capital <= 0:
        return 0.0
    return max(0.0, nopat / invested_capital)


def dynamic_terminal_growth(
    ebit: float,
    tax_rate: float,
    total_assets: float,
    cash: float,
    current_liabilities: float,
    short_term_debt: float | None = None,
    long_term_debt: float | None = None,
    capex: float = 0.0,
    depreciation: float = 0.0,
    change_nwc: float = 0.0,
    cap: float = 0.04,
    floor: float = 0.01,
    discount_rate: float = 0.10,
) -> float:
    def safe_float(x: Any) -> float:
        try:
            val = float(x)
        except (TypeError, ValueError):
            return 0.0
        return 0.0 if np.isnan(val) else val

    ebit = safe_float(ebit)
    tax_rate = safe_float(tax_rate)
    total_assets = safe_float(total_assets)
    cash = safe_float(cash)
    current_liabilities = safe_float(current_liabilities)
    short_term_debt = safe_float(short_term_debt)
    capex = safe_float(capex)
    change_nwc = safe_float(change_nwc)

    nopat = ebit * (1 - tax_rate)
    roic = calculate_roic(ebit, tax_rate, total_assets,
                          cash, current_liabilities, short_term_debt)

    reinvestment = capex + change_nwc
    reinvestment_rate = reinvestment / nopat if nopat > 0 else 0.0

    g_theoretical = roic * reinvestment_rate

    # If destroying value (ROIC < WACC), cap growth more strictly (maturing/melting ice cube)
    actual_cap = cap
    if roic < discount_rate:
        actual_cap = min(cap, 0.02)

    # Enforce a practical floor for terminal growth in healthy companies to avoid overly conservative 1% assumptions
    min_terminal_floor = max(floor, 0.02) if roic >= discount_rate else floor
    dynamic_floor = max(min_terminal_floor, 0.1 * g_theoretical)
    return float(min(max(g_theoretical, dynamic_floor), actual_cap))


def get_ebit_margin(
    income: pd.DataFrame,
    estimates: pd.DataFrame,
    projection_years: int,
    blend_weight: float = 0.7,
) -> np.ndarray:
    """
    Return an array of EBIT margins for each projection year.

    Behaviour:
    - Blends analyst forward EBIT margin (70%) with historical median (30%).
    - For years beyond analyst coverage, linearly converges toward the
      historical median so that margins don't stay frozen at a peak/trough.
    - Floors the entire schedule at 5% to avoid unrealistic negatives.
    """
    revenue_s = income.loc["Total Revenue"].dropna()
    ebit_s = income.loc["Operating Income"].dropna()
    common_idx = revenue_s.index.intersection(ebit_s.index)
    valid_rev = revenue_s.loc[common_idx]
    valid_ebit = ebit_s.loc[common_idx]
    valid_rev = valid_rev[valid_rev > 0]
    valid_ebit = valid_ebit[valid_rev > 0]  # keep only rows with +ve revenue

    if len(valid_rev) == 0:
        return np.full(projection_years, 0.10)

    hist_margins = (valid_ebit / valid_rev).values.astype(float)
    # Use median of last 3 years as anchor for more recent performance, if available
    n_recent = min(len(hist_margins), 3)
    recent_median = np.nanmedian(hist_margins[:n_recent])

    historical_ebit_margin = float(
        0.50 * hist_margins[0] + 0.50 * recent_median
    )
    historical_ebit_margin = float(np.clip(historical_ebit_margin, 0.01, 0.70))

    df = pd.DataFrame(estimates)

    if df.empty or "revenueAvg" not in df.columns or "ebitAvg" not in df.columns:
        # No analyst data: just return the historical margin (constant)
        return np.full(projection_years, max(historical_ebit_margin, 0.05))

    df["date"] = pd.to_datetime(df["date"])
    df["year"] = df["date"].dt.year

    current_year = pd.Timestamp.today().year
    df_future = df[df["year"] >= current_year]
    df_future = df_future[df_future["revenueAvg"].notna()
                          & df_future["ebitAvg"].notna()]

    if df_future.empty:
        return np.full(projection_years, max(historical_ebit_margin, 0.05))

    projected_revenue = df_future["revenueAvg"].astype(float).values
    projected_ebit = df_future["ebitAvg"].astype(float).values
    raw_margin = projected_ebit / \
        np.where(projected_revenue > 0, projected_revenue, np.nan)

    blended_margin = blend_weight * raw_margin + \
        (1 - blend_weight) * historical_ebit_margin
    blended_margin = np.clip(blended_margin, 0.01, 0.70)

    analyst_years = blended_margin[:projection_years].tolist()
    remaining = projection_years - len(analyst_years)

    if remaining > 0:
        # Gently converge analyst margin toward the historical median
        last_margin = float(analyst_years[-1])
        target_margin = historical_ebit_margin
        fade = np.linspace(last_margin, target_margin, remaining + 2)[1:-1]
        analyst_years.extend(fade.tolist())

    ebit_margin = np.array(analyst_years[:projection_years], dtype=float)
    ebit_margin = np.maximum(ebit_margin, 0.05)
    return ebit_margin


def dcf_valuation(
    revenue: pd.Series,
    ebit_margin: Sequence[float] | float,
    tax_rate: float,
    depreciation: pd.Series,
    growth_rates: Sequence[float],
    discount_rate: float,
    terminal_growth: float,
    net_debt: float,
    shares_outstanding: float,
    cashflow: pd.DataFrame,
    balance: pd.DataFrame,
    terminal_roic: float | None = None,
) -> float:
    """
    Core DCF engine. Uses a Sales-to-Capital ratio to project Capex, so that
    Capex scales with incremental revenue rather than with total PPE.
    """

    if discount_rate <= terminal_growth:
        return np.nan

    # --- Input validation ---
    if revenue is None or revenue.dropna().empty:
        raise ValueError("Revenue series is empty.")

    if depreciation is None or depreciation.dropna().empty:
        raise ValueError("Depreciation series is empty.")

    growth_rates = np.asarray(growth_rates, dtype=float)
    if growth_rates.size == 0:
        raise ValueError("growth_rates must be non-empty.")
    if not np.isfinite(growth_rates).all():
        raise ValueError("growth_rates contains non-finite values.")

    if np.isscalar(ebit_margin):
        ebit_margins = np.full(growth_rates.size, float(ebit_margin))
    else:
        ebit_margins = np.asarray(ebit_margin, dtype=float)
        if ebit_margins.shape[0] != growth_rates.size:
            raise ValueError(
                "Length of ebit_margin array must match growth_rates.")

    if shares_outstanding is None or shares_outstanding <= 0:
        raise ValueError("shares_outstanding must be a positive number.")

    if net_debt is None or not np.isfinite(net_debt):
        raise ValueError("net_debt must be a finite numeric value.")

    # --- Last known values ---
    revenue_clean = revenue.dropna().sort_index(ascending=False)
    last_revenue = float(revenue_clean.iloc[0])
    last_depreciation = float(depreciation.dropna().iloc[0])

    if "Capital Expenditure" not in cashflow.index:
        raise ValueError(
            "Missing 'Capital Expenditure' in cash flow statement.")
    capex_series = (-cashflow.loc["Capital Expenditure"]
                    ).dropna().sort_index(ascending=False)
    last_capex = float(capex_series.iloc[0])

    if "Accounts Receivable" not in balance.index or "Accounts Payable" not in balance.index:
        raise ValueError(
            "Balance sheet must contain 'Accounts Receivable' and 'Accounts Payable'.")

    last_ar = float(balance.loc["Accounts Receivable"].iloc[0])
    last_ap = float(balance.loc["Accounts Payable"].iloc[0])

    if "Inventory" in balance.index:
        last_inventory = float(balance.loc["Inventory"].iloc[0])
        if np.isnan(last_inventory):
            last_inventory = 0.0
    else:
        last_inventory = 0.0

    # --- Depreciation rate (as % of revenue) ---
    dep_rate = last_depreciation / last_revenue if last_revenue > 0 else 0.0

    # --- Sales-to-Capital ratio ---
    sales_to_capital = compute_sales_to_capital(
        revenue_clean, capex_series, depreciation.dropna().sort_index(ascending=False),
        balance, last_revenue, last_capex
    )

    # Clamp to reasonable bounds [1.0, 20.0]  # Floor of 1.0 per user request
    sales_to_capital = float(np.clip(sales_to_capital, 1.0, 20.0))

    # Historical capex as % of revenue — used for guardrails
    hist_capex_pct = last_capex / last_revenue if last_revenue > 0 else 0.05

    ufcf_forecast: list[float] = []
    current_ar = last_ar
    current_inventory = last_inventory
    current_ap = last_ap
    current_revenue = last_revenue

    last_nopat = 0.0

    for i, g in enumerate(growth_rates):
        projected_revenue = current_revenue * (1 + g)
        delta_revenue = projected_revenue - current_revenue
        margin = ebit_margins[i]

        projected_ebit = projected_revenue * margin
        projected_nopat = projected_ebit * (1 - tax_rate)
        last_nopat = projected_nopat

        # Depreciation: fixed % of revenue
        projected_depreciation = projected_revenue * dep_rate

        # Net capex via Sales-to-Capital: net_capex = delta_revenue / sales_to_capital
        # Gross capex = net_capex + depreciation
        if delta_revenue >= 0:
            net_capex = delta_revenue / sales_to_capital
        else:
            # Shrinking revenue: no net new investment needed, allow slight contraction
            # will be negative (capex falls)
            net_capex = delta_revenue / sales_to_capital
        projected_capex = net_capex + projected_depreciation

        # Guardrail: projected capex as % of revenue shouldn't exceed 2x or drop below 0.25x historical
        capex_pct = projected_capex / \
            projected_revenue if projected_revenue > 0 else hist_capex_pct
        capex_pct = float(
            np.clip(capex_pct, hist_capex_pct * 0.25, hist_capex_pct * 2.0))
        projected_capex = capex_pct * projected_revenue

        # NWC — ratios off last known revenue (stable working capital assumption)
        ar_ratio = last_ar / last_revenue if last_revenue > 0 else 0.0
        inv_ratio = last_inventory / last_revenue if last_revenue > 0 else 0.0
        ap_ratio = last_ap / last_revenue if last_revenue > 0 else 0.0

        projected_ar = projected_revenue * ar_ratio
        projected_inventory = projected_revenue * inv_ratio
        projected_ap = projected_revenue * ap_ratio

        projected_changenwc = (
            (projected_ar + projected_inventory - projected_ap)
            - (current_ar + current_inventory - current_ap)
        )

        projected_ufcf = (
            projected_nopat
            + projected_depreciation
            - projected_capex
            - projected_changenwc
        )
        ufcf_forecast.append(float(projected_ufcf))

        current_ar = projected_ar
        current_inventory = projected_inventory
        current_ap = projected_ap
        current_revenue = projected_revenue

    discounted_ufcf = [
        ufcf / ((1 + discount_rate) ** (i + 1)) for i, ufcf in enumerate(ufcf_forecast)
    ]

    # --- Terminal Value Calculation ---
    # Unlink Terminal ROIC from WACC if terminal_roic is provided.
    # Textbook: terminal_reinvestment_rate = terminal_growth / terminal_roic
    # If terminal_roic is None, default to terminal_roic = discount_rate (conservative WACC assumption)
    t_roic = terminal_roic if terminal_roic is not None else discount_rate

    # Ensure terminal_roic is at least equal to terminal_growth to avoid >100% reinvestment
    t_roic = max(t_roic, terminal_growth + 0.01)

    terminal_reinvestment_rate = terminal_growth / t_roic
    terminal_nopat = last_nopat * (1 + terminal_growth)
    terminal_ufcf = terminal_nopat * (1 - terminal_reinvestment_rate)

    terminal_value = terminal_ufcf / (
        discount_rate - terminal_growth
    )
    discounted_terminal_value = terminal_value / \
        ((1 + discount_rate) ** len(growth_rates))

    ev = float(sum(discounted_ufcf) + discounted_terminal_value)
    equity_value = ev - float(net_debt)

    implied_share_price = max(0.0, equity_value / float(shares_outstanding))
    return float(implied_share_price)


def project_items(
    revenue: pd.Series,
    ebit_margin: Sequence[float] | float,
    tax_rate: float,
    depreciation: pd.Series,
    growth_rates: Sequence[float],
    cashflow: pd.DataFrame,
    balance: pd.DataFrame,
) -> tuple[list[float], list[float], list[float]]:
    """
    Helper for Excel export: project depreciation, CapEx, and change in NWC.
    Uses the same Sales-to-Capital logic as dcf_valuation.
    """
    revenue_clean = revenue.dropna().sort_index(ascending=False)
    last_revenue = float(revenue_clean.iloc[0])

    dep_series = get_yf_item(cashflow, "depreciation")
    if dep_series is None:
        raise ValueError(
            "Could not locate depreciation in cash flow statement.")
    dep_series_clean = dep_series.dropna().sort_index(ascending=False)
    if dep_series_clean.empty:
        raise ValueError("Depreciation series is empty after dropping NaNs.")
    last_depreciation = float(dep_series_clean.iloc[0])

    if "Capital Expenditure" not in cashflow.index:
        raise ValueError(
            "Missing 'Capital Expenditure' in cash flow statement.")
    capex_series = (-cashflow.loc["Capital Expenditure"]
                    ).dropna().sort_index(ascending=False)
    last_capex = float(capex_series.iloc[0])

    # --- Sales-to-Capital ratio ---
    sales_to_capital = compute_sales_to_capital(
        revenue_clean, capex_series, dep_series_clean, balance, last_revenue, last_capex
    )

    if "Accounts Receivable" not in balance.index or "Accounts Payable" not in balance.index:
        raise ValueError(
            "Balance sheet must contain 'Accounts Receivable' and 'Accounts Payable'.")

    last_ar = float(balance.loc["Accounts Receivable"].iloc[0])
    last_ap = float(balance.loc["Accounts Payable"].iloc[0])

    if "Inventory" in balance.index:
        last_inventory = float(balance.loc["Inventory"].iloc[0])
        if np.isnan(last_inventory):
            last_inventory = 0.0
    else:
        last_inventory = 0.0

    if np.isscalar(ebit_margin):
        ebit_margins = np.full(len(growth_rates), ebit_margin)
    else:
        ebit_margins = np.array(ebit_margin)
        if len(ebit_margins) != len(growth_rates):
            raise ValueError(
                "Length of ebit_margin array must match growth_rates")

    # --- Depreciation rate (% of revenue) ---
    dep_rate = last_depreciation / last_revenue if last_revenue > 0 else 0.0

    sales_to_capital = float(np.clip(sales_to_capital, 1.0, 20.0))
    hist_capex_pct = last_capex / last_revenue if last_revenue > 0 else 0.05

    current_ar = last_ar
    current_inventory = last_inventory
    current_ap = last_ap
    current_revenue = last_revenue

    projected_depreciation_all: list[float] = []
    projected_capex_all: list[float] = []
    projected_changenwc_all: list[float] = []

    for i, g in enumerate(growth_rates):
        projected_revenue = current_revenue * (1 + g)
        delta_revenue = projected_revenue - current_revenue

        # Depreciation
        projected_depreciation = projected_revenue * dep_rate

        # Capex via Sales-to-Capital
        net_capex = delta_revenue / sales_to_capital
        projected_capex = net_capex + projected_depreciation

        # Guardrails
        capex_pct = projected_capex / \
            projected_revenue if projected_revenue > 0 else hist_capex_pct
        capex_pct = float(
            np.clip(capex_pct, hist_capex_pct * 0.25, hist_capex_pct * 2.0))
        projected_capex = capex_pct * projected_revenue

        # NWC
        ar_ratio = last_ar / last_revenue if last_revenue > 0 else 0.0
        inv_ratio = last_inventory / last_revenue if last_revenue > 0 else 0.0
        ap_ratio = last_ap / last_revenue if last_revenue > 0 else 0.0

        projected_ar = projected_revenue * ar_ratio
        projected_inventory = projected_revenue * inv_ratio
        projected_ap = projected_revenue * ap_ratio

        projected_changenwc = (
            (projected_ar + projected_inventory - projected_ap)
            - (current_ar + current_inventory - current_ap)
        )

        projected_depreciation_all.append(float(projected_depreciation))
        projected_capex_all.append(float(projected_capex))
        projected_changenwc_all.append(float(projected_changenwc))

        current_ar = projected_ar
        current_inventory = projected_inventory
        current_ap = projected_ap
        current_revenue = projected_revenue

    return projected_depreciation_all, projected_capex_all, projected_changenwc_all


def fill_excel(
    stock: str,
    ticker: yf.Ticker,
    info: Dict[str, Any],
    income: pd.DataFrame,
    cashflow: pd.DataFrame,
    balance: pd.DataFrame,
    discount_rate: float,
    terminal_growth: float,
    growth_rates: Sequence[float],
    ebit_margin: Sequence[float],
    tax_rate: float,
) -> str:
    """
    Populate the 10-year DCF Excel template (dcf_template_10y.xlsx).

    Layout (unchanged from 5-year template except the forecast extends
    from column J through column S instead of column N):
      Row 13  – Revenue (historical in F-I, forecast in J-S)
      Row 14  – Revenue growth % forecast  (J-S)
      Row 16  – EBIT historical  (F-I)
      Row 17  – EBIT % of sales forecast  (J-S)
      Row 19  – Tax Provision historical  (F-I)
      Row 20  – Tax % of EBIT forecast    (J-S)
      Row 23  – D&A (historical F-I, forecast J-S)
      Row 26  – CapEx (historical F-I, forecast J-S)
      Row 29  – Change in NWC (historical F-I, forecast J-S)
      E9/E10  – WACC / TGR
      S59/S60/S62 – Cash / Debt / Shares (moved from N to S)
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))

    template_path = os.path.join(base_dir, "dcf_template_10y.xlsx")
    output_path = os.path.join(base_dir, f"{stock} DCF.xlsx")

    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb["DCF"]

    short_name = (info.get("shortName") or stock).upper()
    ws["B2"] = f"{short_name} DCF"
    ws["C4"] = stock
    ws["C5"] = f"{datetime.now():%m/%d/%Y}"
    ws["G5"] = ticker.info.get("currentPrice")
    ws["I12"] = income.columns[0].year

    ws["D9"] = discount_rate
    ws["D10"] = terminal_growth
    ws["E9"] = discount_rate
    ws["E10"] = terminal_growth

    # ── Historical financials (columns F=6 → I=9, most-recent first) ────────
    revenue_hist = income.loc["Total Revenue"].dropna()
    ebit_hist = income.loc["Operating Income"]
    past_taxes = income.loc["Tax Provision"]

    for i, value in enumerate(revenue_hist.tolist()):
        ws.cell(row=13, column=9 - i, value=value)
    for i, value in enumerate(ebit_hist.tolist()):
        ws.cell(row=16, column=9 - i, value=value)
    for i, value in enumerate(past_taxes.tolist()):
        ws.cell(row=19, column=9 - i, value=value)

    # ── Forecast assumptions (columns J=10 → S=19, Year 1 → Year 10) ────────
    n_forecast = len(growth_rates)  # should be 10
    for i, value in enumerate(list(growth_rates)):
        ws.cell(row=14, column=10 + i, value=value)
    for i, value in enumerate(list(ebit_margin)):
        ws.cell(row=17, column=10 + i, value=value)
    for col_num in range(10, 10 + n_forecast):
        ws.cell(row=20, column=col_num).value = tax_rate

    # ── D&A, CapEx, Change in NWC ─────────────────────────────────────────────
    depreciation_series = get_yf_item(cashflow, "depreciation")
    if depreciation_series is None:
        raise ValueError(
            "Could not locate depreciation in cash flow statement for Excel export.")

    capex = -cashflow.loc["Capital Expenditure"]

    cash_series = get_yf_item(balance, "cash")
    cash_series = cash_series if cash_series is not None and not cash_series.empty else pd.Series([
                                                                                                  0.0])

    current_debt_series = get_yf_item(balance, "current debt")
    current_debt_series = current_debt_series if current_debt_series is not None and not current_debt_series.empty else pd.Series([
                                                                                                                                  0.0])

    nwc = (balance.loc["Current Assets"] - cash_series) - (
        balance.loc["Current Liabilities"] - current_debt_series
    )
    changenwc_hist = pad_with_mean(nwc - nwc.shift(-1), 4)

    # Historical D&A, CapEx, ΔNWC into columns I→F (most-recent = col 9)
    for i, value in enumerate(depreciation_series.tolist()):
        ws.cell(row=23, column=9 - i, value=value)
    for i, value in enumerate(capex.tolist()):
        ws.cell(row=26, column=9 - i, value=value)
    for i, value in enumerate(changenwc_hist.tolist()):
        ws.cell(row=29, column=9 - i, value=value)

    # Projected D&A, CapEx, ΔNWC into columns J→S
    depreciation_list, capex_list, changenwc_list = project_items(
        revenue=revenue_hist,
        ebit_margin=ebit_margin,
        tax_rate=tax_rate,
        depreciation=depreciation_series,
        growth_rates=growth_rates,
        cashflow=cashflow,
        balance=balance,
    )
    for i, value in enumerate(depreciation_list):
        ws.cell(row=23, column=10 + i, value=value)
    for i, value in enumerate(capex_list):
        ws.cell(row=26, column=10 + i, value=value)
    for i, value in enumerate(changenwc_list):
        ws.cell(row=29, column=10 + i, value=value)

    # ── Balance-sheet plug-ins — now in column S (col 19) ────────────────────
    ws["S59"] = balance.loc["Cash And Cash Equivalents"].iloc[0]
    ws["S60"] = ticker.info.get("totalDebt")
    ws["S62"] = ticker.info.get("sharesOutstanding")

    wb.save(output_path)
    return output_path


def fill_excel_cca(
    stock: str,
    info: Dict[str, Any],
    peer_data: Sequence[Dict[str, Any]], income, API_KEY, weights
) -> str:
    """
    Populate the CCA Excel template with peer data and target financials.
    Only fills 'input' fields, leaving formula cells untouched.
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, "cca_template.xlsx")
    output_path = os.path.join(base_dir, f"{stock} CCA.xlsx")

    shutil.copy(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb["CCA"]

    # Target Company Header
    short_name = (info.get("shortName") or stock).upper()
    ws["B2"] = f"{short_name} Comparable Companies Analysis"
    ws["B3"] = f"{datetime.now():%m/%d/%Y}"

    # fill in target company data
    ws["B9"] = info.get("shortName")
    ws["C9"] = stock
    ws["D9"] = info.get("currentPrice")
    ws["B21"] = info.get("longName")

    ws["F9"] = info.get("totalRevenue")
    ws["G9"] = info.get("ebitda")

    # Target NI extraction (Sync with cca.py robust logic)
    target_ni = info.get("netIncome") or info.get("netIncomeToCommon") or (safe_iloc0(
        income.loc["Net Income"]) if (not income.empty and "Net Income" in income.index) else None)

    ws["H9"] = target_ni

    target_pe = info.get("trailingPE") or info.get("forwardPE")

    ws["J9"] = info.get("enterpriseToRevenue")
    ws["K9"] = info.get("enterpriseToEbitda")
    ws["L9"] = target_pe
    ws["M9"] = info.get("trailingPegRatio") or info.get("pegRatio")

    # EPS Growth safety logic
    eps_growth = None
    fmp_growth = pull_info("financial-growth", stock, API_KEY)
    if fmp_growth is not None and isinstance(fmp_growth, list) and len(fmp_growth) > 0:
        eps_growth = fmp_growth[0].get("epsgrowth")
    if eps_growth is None:
        eps_growth = info.get("earningsGrowth")

    ws["N9"] = eps_growth
    ws["P9"] = info.get("returnOnEquity")

    # Adjusted P/E safety check
    adj_denominator = (1 + eps_growth) if eps_growth is not None else None
    if target_pe and adj_denominator and adj_denominator != 0:
        ws["Q9"] = target_pe / adj_denominator
    else:
        ws["Q9"] = None

    # Peer Data Rows populate
    start_row = 6
    for i, peer in enumerate(peer_data):
        row = start_row + i
        if row >= 9:
            row += 1  # Skip the target's row (9)
        ws.cell(row=row, column=2, value=peer.get("name"))
        ws.cell(row=row, column=3, value=peer.get("ticker"))
        ws.cell(row=row, column=4, value=peer.get("price"))
        ws.cell(row=row, column=6, value=peer.get("revenue"))
        ws.cell(row=row, column=7, value=peer.get("ebitda"))
        ws.cell(row=row, column=8, value=peer.get("net_income"))
        ws.cell(row=row, column=10, value=peer.get("ev_rev"))
        ws.cell(row=row, column=11, value=peer.get("ev_ebitda"))
        ws.cell(row=row, column=12, value=peer.get("pe"))
        ws.cell(row=row, column=13, value=peer.get("peg"))
        ws.cell(row=row, column=14, value=peer.get("eps_growth"))
        ws.cell(row=row, column=16, value=peer.get("roe"))
        ws.cell(row=row, column=17, value=peer.get("adj_pe"))

    # Populate Net Debt and Shares (Search column B for labels)
    net_debt_val = (info.get("totalDebt") or 0) - (info.get("totalCash") or 0)
    shares_val = info.get("sharesOutstanding")

    ws["J23"] = net_debt_val
    ws["J25"] = shares_val

    # Implied Weights
    ws["J28"] = weights[0]
    ws["K28"] = weights[1]
    ws["L28"] = weights[2]

    wb.save(output_path)
    return output_path


def solve_for_revenue_growth(
    current_price: float,
    revenue: pd.Series,
    ebit_margin: Sequence[float],
    tax_rate: float,
    depreciation: pd.Series,
    discount_rate: float,
    terminal_growth: float,
    net_debt: float,
    shares_outstanding: float,
    cashflow: pd.DataFrame,
    balance: pd.DataFrame,
    terminal_roic: float | None = None,
    tolerance: float = 0.05,
    max_iter: int = 50,
) -> float:
    """
    Reverse DCF: Solves for the constant ANNUAL revenue growth rate 
    that yields the given current_price.
    Uses a simple bisection method.
    """
    low, high = -0.5, 2.0  # Search between -50% and +200% annual growth

    # Check if price increases with growth
    p_low = dcf_valuation(revenue, ebit_margin, tax_rate, depreciation, np.full(len(ebit_margin), low), discount_rate,
                          terminal_growth, net_debt, shares_outstanding, cashflow, balance, terminal_roic=terminal_roic)
    p_high = dcf_valuation(revenue, ebit_margin, tax_rate, depreciation, np.full(len(ebit_margin), high), discount_rate,
                           terminal_growth, net_debt, shares_outstanding, cashflow, balance, terminal_roic=terminal_roic)

    # If the price at max growth is lower than at min growth, growth is value-destructive
    if p_high < p_low:
        return -999.0

    for _ in range(max_iter):
        mid = (low + high) / 2
        growth_rates = np.full(len(ebit_margin), mid)

        try:
            implied_price = dcf_valuation(
                revenue=revenue,
                ebit_margin=ebit_margin,
                tax_rate=tax_rate,
                depreciation=depreciation,
                growth_rates=growth_rates,
                discount_rate=discount_rate,
                terminal_growth=terminal_growth,
                net_debt=net_debt,
                shares_outstanding=shares_outstanding,
                cashflow=cashflow,
                balance=balance,
                terminal_roic=terminal_roic,
            )
        except Exception:
            return np.nan

        if abs(implied_price - current_price) < tolerance:
            return mid

        if implied_price > current_price:
            high = mid
        else:
            low = mid

    return (low + high) / 2


def solve_for_tgr(
    current_price: float,
    revenue: pd.Series,
    ebit_margin: Sequence[float],
    tax_rate: float,
    depreciation: pd.Series,
    growth_rates: Sequence[float],
    discount_rate: float,
    net_debt: float,
    shares_outstanding: float,
    cashflow: pd.DataFrame,
    balance: pd.DataFrame,
    terminal_roic: float | None = None,
    tolerance: float = 0.05,
    max_iter: int = 50,
) -> float:
    """
    Reverse DCF: Solves for the terminal growth rate (TGR) 
    that yields the given current_price.
    Uses a simple bisection method.
    """
    # TGR must be less than discount_rate.Search between -10% and (discount_rate - 0.0001)
    low, high = -0.10, discount_rate - 0.0001

    for _ in range(max_iter):
        mid = (low + high) / 2

        try:
            implied_price = dcf_valuation(
                revenue=revenue,
                ebit_margin=ebit_margin,
                tax_rate=tax_rate,
                depreciation=depreciation,
                growth_rates=growth_rates,
                discount_rate=discount_rate,
                terminal_growth=mid,
                net_debt=net_debt,
                shares_outstanding=shares_outstanding,
                cashflow=cashflow,
                balance=balance,
                terminal_roic=terminal_roic,
            )
        except Exception:
            return np.nan

        if abs(implied_price - current_price) < tolerance:
            return mid

        if implied_price > current_price:
            high = mid
        else:
            low = mid

    return (low + high) / 2
